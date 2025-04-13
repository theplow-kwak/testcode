#!/usr/bin/env python3

import os
import subprocess
import csv
import re
from openpyxl import load_workbook
import logging

# 로깅 설정
logging.basicConfig(level=logging.INFO)

# 상수 변수 설정
CURRENT_PATH = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = CURRENT_PATH
DATA_PATH = f"{CURRENT_PATH}/data"
LOG_PATH = f"{CURRENT_PATH}/logs"
LOG_ERROR = logging.ERROR
LOG_INFO = logging.INFO


class CommandRunner:
    def VmExec(self, cmd: str, ignoreError: bool = True) -> str:
        try:
            logging.info(f"Running command: {cmd}")
            result = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, check=not ignoreError)
            if result.returncode != 0:
                logging.warning(f"Command failed but ignored. Error: {result.stdout.strip()}")
            else:
                logging.info(f"Command executed successfully. Output: {result.stdout.strip()}")
            return result.stdout.strip()
        except subprocess.CalledProcessError as e:
            logging.error(f"An error occurred while executing the command: {e.stderr}")
            if not ignoreError:
                raise e
        return ""


class nvme_common:
    def __init__(self, hostRoot, ostype="Linux"):
        self.status = {}  # Dictionary that keeps status such as fna,nsze
        self.hostRoot = hostRoot if hostRoot else CommandRunner()
        self.ostype = ostype

    def run(self, target_dev, opt_main, opt_sub="", ignoreError=False):
        sudo = "" if os.geteuid() == 0 else "sudo "
        cmd = f"{sudo}nvme {opt_main.strip()} {target_dev.strip()} {opt_sub.strip()}"
        return self.hostRoot.VmExec(cmd, ignoreError)

    def get_log(self, target_dev, log_id, namespace="", length=""):
        opt_main = "get-log"
        opt_sub = f"-i {log_id} -l {length}"
        if namespace:
            opt_sub += f" -n {namespace}"
        return self.run(target_dev, opt_main, opt_sub, ignoreError=True)

    def get_device_id(self, device):
        device_id_match = re.search(r"nvme\d+", device)
        if device_id_match:
            device = device_id_match.group()
            return self.hostRoot.VmExec(f"cat /sys/class/nvme/{device}/device/device")
        return None

    def get_fw_version(self, target_dev):
        opt_main = "ocp get-fw-rev"
        opt_sub = ""
        return self.run(target_dev, opt_main, opt_sub, ignoreError=True)


class ExcelData:
    def __init__(self, path, start_row=1):
        self.path = path
        self.start_row = start_row
        self.data = self.read_excel_data()

    def read_excel_data(self):
        workbook = load_workbook(self.path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[self.start_row]]
        data = []
        for row in sheet.iter_rows(min_row=self.start_row + 1, values_only=True):
            if not row[0]:
                break
            item = {headers[i]: value for i, value in enumerate(row)}
            data.append(item)
        workbook.close()
        return data


class SmartData:
    def __init__(self, hostRoot, ostype="Linux", target_device=None):
        self.smart_before_fname = os.path.join(LOG_PATH, "Stat_SMARTExtension_Before.csv")
        self.smart_after_fname = os.path.join(LOG_PATH, "Stat_SMARTExtension_After.csv")
        self.smart_summary_fname = os.path.join(LOG_PATH, "Stat_SMARTExtension_Summary.csv")
        config_file = os.path.join(CONFIG_PATH, "smart", "SMART_Version_Management.xlsx")
        self.smart_criteria = ExcelData(config_file, 3).data
        self.hostRoot = hostRoot
        self.fail_count = 0
        self.warning_count = 0
        self.nvme = nvme_common(self.hostRoot, "Linux")
        if target_device:
            self.target_device = target_device

    def read_csv(self, file_path):
        with open(file_path, mode="r", newline="", encoding="utf-8") as file:
            return list(csv.DictReader(file))

    def get_customer_code(self):
        internal_firmware_revision = self.get_internal_firmware_revision()
        if len(internal_firmware_revision) == 8:
            customer_code = {"H": "HP", "M": "MS", "3": "LENOVO", "D": "DELL", "5": "ASUS", "K": "FACEBOOK"}.get(internal_firmware_revision[4], "UNKNOWN")
            return customer_code
        return "UNKNOWN"

    def get_internal_firmware_revision(self):
        fw_version = self.nvme.get_fw_version(self.target_device)
        fw_version_match = re.search(r"Internal Firmware Revision: ([A-Za-z0-9]+)", fw_version)
        if fw_version_match:
            return fw_version_match.group(1)
        return ""

    def get_project_code(self):
        device_string = self.nvme.get_device_id(self.target_device)
        dev_part_match = re.search(r"0x([0-9a-fA-F]+)", device_string)
        if dev_part_match:
            dev_part_value = dev_part_match.group(1)
            project_code = next((key for key in self.smart_criteria[0].keys() if key is not None and re.search(dev_part_value, key, re.IGNORECASE)), None)
        if not project_code:
            project_code = next((key for key in self.smart_criteria[0].keys() if key is not None and re.search("SSD01", key, re.IGNORECASE)), None)
        return project_code

    def evaluate_condition(self, value_before, value_after, conditions):
        for condition in conditions.split(";"):
            if match := re.match(r"^(\w+):(\d+)$", condition):
                operator = match.group(1)
                threshold = float(match.group(2))
                if (
                    (operator == "gt" and (value_before > threshold or value_after > threshold))
                    or (operator == "lt" and (value_before < threshold or value_after < threshold))
                    or (operator == "ge" and (value_before >= threshold or value_after >= threshold))
                    or (operator == "le" and (value_before <= threshold or value_after <= threshold))
                    or (operator == "ne" and (value_before != threshold or value_after != threshold))
                    or (operator == "eq" and (value_before == threshold or value_after == threshold))
                ):
                    return True
            else:
                if condition == "inc" and value_after > value_before:
                    return True
                elif condition == "dec" and value_after < value_before:
                    return True
                elif condition == "ne" and value_after != value_before:
                    return True
        return False

    def compare_smart_customer(self, customer_code, project_code):
        smart_before = [item for item in self.smart_before if item["customer"] == customer_code]
        smart_after = [item for item in self.smart_after if item["customer"] == customer_code]

        for criteria in filter(lambda x: x["customer"] == customer_code, self.smart_criteria):
            byte_offset = criteria["byte_offset"]
            description = criteria["field_name"]
            condition = criteria.get(project_code, "")
            value_before_obj = next((item for item in smart_before if item["byte_offset"] == byte_offset), None)
            value_after_obj = next((item for item in smart_after if item["byte_offset"] == byte_offset), None)
            result = "Not Evaluated"

            if condition:
                value_before = float(value_before_obj["value"]) if value_before_obj else 0.0
                value_after = float(value_after_obj["value"]) if value_after_obj else 0.0
                result = criteria["criteria"] if self.evaluate_condition(value_before, value_after, condition) else "PASS"
                if result.upper().strip() == "WARNING":
                    self.warning_count += 1
                    logging.warning(f"SMART:WARNING - {customer_code},{byte_offset},{description},{condition}: {value_before} <==> {value_after}")
                elif result.upper().strip() == "FAIL":
                    self.fail_count += 1
                    logging.error(f"SMART:FAIL - {customer_code},{byte_offset},{description},{condition}: {value_before} <==> {value_after}")
            if value_before_obj and value_after_obj:
                self.csv_writer.writerow([customer_code, byte_offset, value_before_obj["hex_value"], value_after_obj["hex_value"], description, result])
            else:
                self.csv_writer.writerow([customer_code, byte_offset, "", "", description, result])

    def compare_smart_data(self):
        self.smart_before = self.read_csv(self.smart_before_fname)
        if not os.path.isfile(self.smart_after_fname):
            self.save_smart_data(step="after")
        self.smart_after = self.read_csv(self.smart_after_fname)
        summary_file = open(self.smart_summary_fname, "w")
        self.csv_writer = csv.writer(summary_file)
        self.csv_writer.writerow(["customer", "byte_offset", "Before Value (HEX)", "After Value (HEX)", "field_name", "result"])

        customer_code = self.get_customer_code()
        project_code = self.get_project_code()
        nvme_customer_code = {"HP": "NVME_HP", "DELL": "NVME_DELL"}.get(customer_code, "NVME_GEN")

        self.compare_smart_customer("NVME", project_code)
        self.compare_smart_customer(nvme_customer_code, project_code)
        self.compare_smart_customer(customer_code, project_code)
        self.compare_smart_customer("WAI", project_code)
        self.compare_smart_customer("WAF", project_code)

        summary_file.close()
        if self.fail_count > 0:
            logging.error(f"[SMART][{customer_code}][{project_code}] Test Failed - Warning: {self.warning_count}, Fail: {self.fail_count}")
            return False
        else:
            logging.info(f"[SMART][{customer_code}][{project_code}] Test Passed - Warning: {self.warning_count}, Fail: {self.fail_count}")
            return True

    def get_bytes(self, data, byte_offset):
        if match := re.match(r"^(\d+):(\d+)$", byte_offset):
            end = int(match.group(1))
            start = int(match.group(2))
            return data[start : end + 1]
        elif match := re.match(r"^(\d+)$", byte_offset):
            start = int(match.group(1))
            return [data[start]]
        else:
            raise ValueError(f"Invalid byteOffset format: {byte_offset}")

    def bytes_to_number(self, bytes_data):
        if not bytes_data:
            return 0
        bytes_data = bytes.fromhex("".join(bytes_data))
        return int.from_bytes(bytes_data, byteorder="little")

    def bytes_to_hex_string(self, bytes_data):
        hex_string = []
        size = len(bytes_data)
        if size > 40:
            return ""
        if size in [2, 4, 8, 16]:
            bytes_data = bytes_data[::-1]  # Reverse the byte order for little-endian representation
        for i in range(size):
            hex_string.append(f"{bytes_data[i]}")
        return re.sub(r"(.{8})(?!$)", r"\1 ", "".join(hex_string)).upper()

    def bytes_to_decimal_string(self, bytes_data):
        decimal = "0"
        for byte in reversed(bytes_data):
            carry = int(byte, base=16)
            for i in range(len(decimal) - 1, -1, -1):
                num = (ord(decimal[i]) - ord("0")) * 256 + carry
                decimal = decimal[:i] + chr(ord("0") + num % 10) + decimal[i + 1 :]
                carry = num // 10
            while carry > 0:
                decimal = chr(ord("0") + carry % 10) + decimal
                carry //= 10
        return decimal

    def get_log_page(self, customer_code):
        log_id, log_length = {
            "SSD": ["0xFE", 8192],
            "HP": ["0xC7", 512],
            "MS": ["0xC0", 512],
            "LENOVO": ["0xDF", 512],
            "DELL": ["0xCA", 512],
        }.get(customer_code, ["0x02", 512])

        raw_data = self.nvme.get_log(self.target_device, log_id=log_id, length=log_length)
        if re.match(r"NVMe status: Invalid", raw_data, re.IGNORECASE):
            logging.info(f"Can't get SMART log for {customer_code}({log_id})\n")
            smart_data = ["00"] * log_length
        else:
            smart_data = []
            for line in raw_data.split("\n")[2:]:
                smart_data.extend(line.split(" ")[1:-1])
        return smart_data

    def parse_smart_data(self, smart_data, customer_code):
        for template in filter(lambda x: x["customer"] == customer_code, self.smart_criteria):
            byte_offset = str(template["byte_offset"])
            field_name = template["field_name"]
            bytes_data = self.get_bytes(smart_data, byte_offset)
            if field_name == "Reserved":
                value = ""
                hex_value = ""
            else:
                hex_value = self.bytes_to_hex_string(bytes_data)
                value = self.bytes_to_decimal_string(bytes_data)
            self.csv_writer.writerow([customer_code, byte_offset, hex_value, value, field_name])

    def get_wai_waf(self, smart_data):
        ec_slc_total = self.bytes_to_number(self.get_bytes(smart_data, "227:224"))
        ec_tlc_total = self.bytes_to_number(self.get_bytes(smart_data, "243:240"))
        written_to_slc_user = self.bytes_to_number(self.get_bytes(smart_data, "127:120"))
        written_to_tlc_user = self.bytes_to_number(self.get_bytes(smart_data, "119:112"))
        write_from_host = self.bytes_to_number(self.get_bytes(smart_data, "111:104"))
        host_writes = write_from_host or 1  # Prevent division by zero
        wai = (written_to_slc_user + written_to_tlc_user) / host_writes
        waf = (ec_slc_total + ec_tlc_total) / host_writes
        return wai, waf

    def save_smart_data(self, step="before"):
        if self.target_device is None:
            logging.info("Target device not specified")
            return
        filename = self.smart_before_fname if step == "before" else self.smart_after_fname
        with open(filename, "w", newline="") as smart_file:
            self.csv_writer = csv.writer(smart_file)
            self.csv_writer.writerow(["customer", "byte_offset", "hex_value", "value", "field_name"])

            nvme_customer_code = {"HP": "NVME_HP", "DELL": "NVME_DELL"}.get(self.get_customer_code(), "NVME_GEN")
            smart_data = self.get_log_page("NVME")
            self.parse_smart_data(smart_data, "NVME")
            self.parse_smart_data(smart_data, nvme_customer_code)

            customer_codes = ["HP", "MS", "LENOVO", "DELL", "SKH"]
            for code in customer_codes:
                smart_data = self.get_log_page(code)
                self.parse_smart_data(smart_data, code)

            wai, waf = self.get_wai_waf(smart_data[512:])
            self.csv_writer.writerow(["WAI", "000", f"{wai:.6f}", f"{wai:.6f}", "[ETC] (ec_slc_total + ec_tlc_total) / write_from_host"])
            self.csv_writer.writerow(["WAF", "000", f"{waf:.6f}", f"{waf:.6f}", "[ETC] (written_to_tlc_user + written_to_slc_buf) / write_from_host"])


# 사용 예시
if __name__ == "__main__":
    smart_comparer = SmartData("", target_device="/dev/nvme0")
    smart_comparer.save_smart_data()
    smart_comparer.compare_smart_data()
